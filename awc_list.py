import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


def create_rpf_structure_xlsx(root_dir: str, xlsx_path: str):
    """
    遍历root_dir下所有以.rpf结尾的文件夹及其子文件夹，生成展示包含关系的Excel表格

    参数:
        root_dir: 要遍历的根目录
        xlsx_path: 生成的Excel文件路径
    """
    # 收集所有.rpf文件夹和它们的子文件夹
    rpf_folders = []
    subfolders_dict = {}  # {rpf_path: [subfolder1, subfolder2, ...]}

    # 遍历root_dir寻找.rpf结尾的文件夹
    for entry in os.listdir(root_dir):
        full_path = os.path.join(root_dir, entry)
        if os.path.isdir(full_path) and entry.endswith('.rpf'):
            rpf_folders.append(full_path)

            # 收集第一层子文件夹
            subfolders = []
            for sub_entry in os.listdir(full_path):
                sub_full_path = os.path.join(full_path, sub_entry)
                if os.path.isdir(sub_full_path):
                    subfolders.append(sub_entry)

            subfolders_dict[full_path] = subfolders

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "RPF Structure"

    # 设置表头
    ws['A1'] = "RPF Folders"
    ws['B1'] = "Subfolders"

    # 写入数据并准备合并单元格
    current_row = 2
    merge_ranges = []

    for rpf_path in rpf_folders:
        subfolders = subfolders_dict[rpf_path]
        rpf_name = os.path.basename(rpf_path)

        if not subfolders:
            # 如果没有子文件夹，只写入RPF名称
            ws.cell(row=current_row, column=1, value=rpf_name)
            current_row += 1
        else:
            # 写入RPF名称并设置合并范围
            start_row = current_row
            end_row = current_row + len(subfolders) - 1
            ws.cell(row=current_row, column=1, value=rpf_name)

            if start_row != end_row:
                merge_ranges.append(f"A{start_row}:A{end_row}")

            # 写入子文件夹
            for subfolder in subfolders:
                ws.cell(row=current_row, column=2, value=subfolder)
                current_row += 1

    # 应用合并单元格
    for range_string in merge_ranges:
        ws.merge_cells(range_string)

    # 设置垂直居中和自动换行
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 调整列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40

    # 保存Excel文件
    wb.save(xlsx_path)

# 使用示例
if __name__ == '__main__':
    create_rpf_structure_xlsx('E:\\ai\\GTA5_Chinese\\friends', './in_xlsx/friends_list.xlsx')