import os
import os
import re
from time import sleep

import openpyxl
import pandas as pd
import py7zr
from openai import OpenAI
from profanity_check import predict

in_xlsx_root = 'in_xlsx'
ori_oxt_root = 'subtitles/cnsim'
out_dir = 'out_oxt'
in_audio_root = 'in_audio'
out_sfx_root = 'out_sfx'
common_punctuation = ['.', '!', '?', '。', '！', '？', '，', '、', '：', '；', '‘', '’', '“', '”', '"', "'", '(', ')', '[',
                      ']', '{', '}', '<', '>', '—', '–', '…', '—', '‘', '’', '“', '”', '《', '》', '【', '】', '『', '』',
                      '﹃', '﹄', '〔', '〕']


def natural_sort_key(s):
    """
    生成自然排序的key函数
    """
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s)]


def censor_bad_words(sentence):
    """
    使用 profanity-check 库检测并屏蔽脏话。
    将脏话单词的每个字母之间插入 "*"。
    """
    words = sentence.split()

    for i in range(len(words)):
        if predict([words[i]])[0]:
            censored_word = '*'.join(words[i])
            words[i] = censored_word

    censored_sentence = ' '.join(words)

    return censored_sentence


# 中文检测
def is_chinese(text):
    for char in text:
        if '\u4e00' <= char <= '\u9fff':
            return True
    return False


# 从oxt文件中读取字典
def read_oxt(oxt_file, is_cn=True, audio_prefix=None):
    ret = {}
    sub_prefix = ['~z~', '~t~']
    audio_prefix = [os.path.basename(oxt_file).split('aud')[0].upper()] if audio_prefix is None else audio_prefix

    def is_audio(value):
        for prefix in audio_prefix:
            if value.startswith(prefix):
                return True
        return False

    def starts_with_sub_prefix(value):
        for prefix in sub_prefix:
            if value.startswith(prefix):
                return True
        return False

    with open(oxt_file, 'rb') as f:
        data = f.read().decode('utf-8')
    for line in data.split('\n'):
        if ' = ' in line:
            key, value = line.split(' = ')

            key = key.strip().replace('\t', '').replace('\n', '').replace('\r', '')
            value = value.strip().replace('\t', '').replace('\n', '').replace('\r', '')
            if starts_with_sub_prefix(value) and (is_chinese(value) and is_cn or not is_cn):  # 记录了字幕的项
                ret[key] = value[3:]
            elif is_audio(value):  # 记录了音频文件的项
                ret[key] = value

    return ret


def read_map_from_xlsx(xlsx_file, key='原台词简中', value='中配台词', filter_role=None):
    """读取xlsx文件，找到'原台词'和'中配台词'的对应关系"""

    # 读取整个Excel文件（包括所有行和列），不指定表头
    df = pd.read_excel(xlsx_file, header=None)

    # 初始化变量来存储找到的行和列索引
    key_col, value_col, start_row = None, None, None

    # 遍历每一行，找到包含“原台词”和“中配台词”标签的行和列
    for i, row in df.iterrows():
        for j, cell in enumerate(row):
            if cell == key:
                key_col = j
                start_row = i
            elif cell == value:
                value_col = j
                start_row = i

        # 一旦找到了两个标签所在的列，就可以退出循环
        if key_col is not None and value_col is not None:
            break

    # 如果找不到所需的列，抛出异常
    if key_col is None or value_col is None:
        raise ValueError("无法找到包含‘key’或‘value’的列，请检查文件格式。")

    # 定义一个过滤函数，检查行中是否包含指定的filter_role
    def row_filter(row):
        # 不是过滤模式
        if filter_role is None or filter_role == '':
            return True
        # 如果row的第一个元素（音频文件名）不存在大写字母，则也过滤掉（过场动画）
        if not any(c.isupper() for c in row.values[0]):
            return False
        return filter_role in row.values

    # 从 start_row + 1 行开始，过滤并提取数据
    filtered_rows = df.iloc[start_row + 1:].apply(row_filter, axis=1)
    filtered_data = df.iloc[start_row + 1:][filtered_rows]

    # 提取过滤后的 key 和 value 列数据
    original_data = filtered_data.iloc[:, key_col].str.strip()
    dub_data = filtered_data.iloc[:, value_col].str.strip()

    # 去除空值，生成键值对映射
    kv_map = dict(zip(original_data, dub_data))

    for k, v in kv_map.items():
        # 如果v是nan，则用k代替
        if v is None or v == '' or pd.isna(v) or v.__len__() == 0:
            kv_map[k] = k
            v = k
        # 如果v中存在一对中英文括号，去掉括号及其内容，例如“AAAA（BBBB）CC(DD)CC” -> "AAAACCCC"
        v = re.sub(r'[$\（][^)\）]*[$\）]', '', v)

        # 如果value的末尾没有中文或英文标点，则添加一个句号
        if v and v[-1] not in common_punctuation:
            kv_map[k] = v + '。'

    return kv_map


class Chatbot:
    def __init__(self, api_key, base_url, engine, sys_prompt):
        self.client = OpenAI(
            api_key=api_key,
            base_url=base_url,
        )
        self.engine = engine
        self.sys_prompt = sys_prompt

    def optimize(self, last_script, cur_script, next_script, temp=0.9):
        query = f"上一句: {last_script}\n当前句: {cur_script}\n下一句: {next_script}\n请直接输出润色后的当前句的台词(简体中文):"

        retry = 0
        while retry < 10:
            try:
                completion = self.client.chat.completions.create(
                    model=self.engine,
                    messages=[
                        {"role": "system", "content": self.sys_prompt},
                        {"role": "user", "content": query},
                    ],
                    temperature=temp)
            except Exception as e:
                print(f"优化失败，重试{retry + 1}次。 " + str(e))
                retry += 1
                sleep(4)
                continue
            break
        if retry >= 10:
            print("优化失败，请检查问题！")
            return None

        return completion.choices[0].message.content

    def translate(self, group, last_sentence, sentence, next_sentence, temp=0.9):
        query = f"语音组: {group}\n上一句: {last_sentence}\n当前句: {sentence}\n下一句: {next_sentence}\n请直接回复当前句翻译后的台词(简体中文):"

        retry = 0
        while retry < 10:
            try:
                completion = self.client.chat.completions.create(
                    model=self.engine,
                    messages=[
                        {"role": "system", "content": self.sys_prompt},
                        {"role": "user", "content": query},
                    ],
                    temperature=temp)
            except Exception as e:
                print(f"翻译失败，重试{retry + 1}次。 " + str(e))
                retry += 1
                sleep(4)
                continue
            break
        if retry >= 10:
            print("翻译失败，请检查问题！")
            return None

        return completion.choices[0].message.content


def read_audio_script_from_xlsx(xlsx_path, audio_col='音频文件', script_col='AI翻译简中'):
    """读取xlsx文件，找到音频文件和台词的对应关系, 返回字典
    Args:
        xlsx_path: xlsx文件路径
        audio_col: 音频文件列名(默认'音频文件')
        script_col: 台词文本列名(默认'AI翻译简中')
    Returns:
        dict: {音频文件名: 对应台词} 的字典
    Raises:
        ValueError: 如果找不到指定列名
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    # 初始化变量
    header_row = None
    audio_col_idx = None
    script_col_idx = None

    # 查找包含两个列名的行作为表头行
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        for cell in row:
            if cell.value == audio_col:
                audio_col_idx = cell.column
            elif cell.value == script_col:
                script_col_idx = cell.column

        # 如果两个列都找到了，记录当前行号并停止搜索
        if audio_col_idx and script_col_idx:
            header_row = row_idx
            break

    # 检查是否找到列
    if not (audio_col_idx and script_col_idx):
        raise ValueError(f"未找到指定列: {audio_col} 或 {script_col}")

    # 收集数据
    result = {}
    for row in sheet.iter_rows(min_row=header_row + 1):  # 从表头行下一行开始读取
        audio_cell = row[audio_col_idx - 1]  # 列索引从1开始，列表从0开始
        script_cell = row[script_col_idx - 1]

        # 确保都有值且不是空字符串
        if audio_cell.value and script_cell.value and str(audio_cell.value).strip() and str(script_cell.value).strip():
            result[str(audio_cell.value).strip()] = str(script_cell.value).strip()

    return result



if __name__ == '__main__':
    xlsx_path = 'in_xlsx/01_0_gtaiv_legacy_support.xlsx'
    audio_col = '音频文件'
    script_col = 'AI翻译简中'
    result = read_audio_script_from_xlsx(xlsx_path, audio_col, script_col)
    print(result)
